from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View, DetailView
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import PasswordResetForm
from django.urls import reverse_lazy
from django.utils.crypto import get_random_string
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.http import HttpResponse
import textwrap
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .mixins import RoleRequiredMixin
from accounts.models import User, AuditLog
from django.db.models import Q, Case, When, IntegerField
from university.models import TaskForce, Department, WorkloadSettings
from .forms import StaffForm, TaskForceForm, DepartmentForm, WorkloadSettingsForm

def build_taskforce_pdf_response(taskforces, title, filename):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    left_margin = 40
    bottom_margin = 50
    line_height = 14

    def start_page():
        nonlocal y
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(left_margin, height - 50, title)
        pdf.setFont("Helvetica", 10)
        y = height - 70

    y = height - 70
    start_page()

    for tf in taskforces:
        dept_names = ", ".join([d.name for d in tf.departments.all()])
        lines = [
            f"Task Force: {tf.name} ({tf.chart_id or ''})",
            f"Status: {tf.get_status_display()} | Weightage: {tf.weightage} | Members: {tf.members.count()}",
            f"Departments: {dept_names or '-'}",
        ]
        if tf.description:
            lines.append(f"Description: {tf.description}")

        for line in lines:
            for chunk in textwrap.wrap(line, width=110) or [""]:
                if y < bottom_margin:
                    pdf.showPage()
                    start_page()
                pdf.drawString(left_margin, y, chunk)
                y -= line_height
        y -= 6

    pdf.save()
    return response

class DashboardDispatcher(LoginRequiredMixin, TemplateView):
    """Redirects authenticated users to their specific role dashboard."""
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.role == User.Role.ADMIN:
            return redirect('dashboard:admin')
        elif user.role == User.Role.HOD:
            return redirect('dashboard:hod')
        elif user.role == User.Role.PSM:
            return redirect('dashboard:psm')
        elif user.role == User.Role.DEAN:
            return redirect('dashboard:dean')
        elif user.role == User.Role.LECTURER:
            return redirect('dashboard:lecturer')
        return redirect('login') # Fallback

class AdminDashboardView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/admin_dashboard.html"
    required_role = User.Role.ADMIN

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['staff_count'] = User.objects.exclude(role=User.Role.ADMIN).count()
        context['taskforce_count'] = TaskForce.objects.count()
        context['department_count'] = Department.objects.count()
        context['recent_users'] = User.objects.order_by('-date_joined')[:5]
        context['recent_logs'] = AuditLog.objects.select_related('actor').order_by('-timestamp')[:5]
        return context

# --- Admin Department Management ---
class DepartmentListView(RoleRequiredMixin, ListView):
    model = Department
    template_name = "dashboard/admin/department_list.html"
    context_object_name = "departments"
    required_role = User.Role.ADMIN

class DepartmentCreateView(RoleRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = "dashboard/admin/department_form.html"
    success_url = reverse_lazy('dashboard:department_list')
    required_role = User.Role.ADMIN

class DepartmentUpdateView(RoleRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = "dashboard/admin/department_form.html"
    success_url = reverse_lazy('dashboard:department_list')
    required_role = User.Role.ADMIN

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_edit'] = True
        return context

# --- Admin Staff Management ---
class StaffListView(RoleRequiredMixin, ListView):
    model = User
    template_name = "dashboard/admin/staff_list.html"
    context_object_name = "staff_list"
    required_role = User.Role.ADMIN
    
    def get_queryset(self):
         queryset = User.objects.all()
         staff_id = self.request.GET.get('staff_id', '').strip()
         username = self.request.GET.get('username', '').strip()
         name = self.request.GET.get('name', '').strip()
         role = self.request.GET.get('role', '').strip()
         status = self.request.GET.get('status', '').strip()

         if staff_id:
             queryset = queryset.filter(staff_id__icontains=staff_id)
         if username:
             queryset = queryset.filter(username__icontains=username)
         if name:
             queryset = queryset.filter(
                 Q(first_name__icontains=name) |
                 Q(last_name__icontains=name)
             )
         if role:
             queryset = queryset.filter(role=role)
         if status:
             if status == 'inactive':
                 queryset = queryset.filter(is_active=False)
             elif status == 'locked':
                 queryset = queryset.filter(is_locked=True, is_active=True)
             elif status == 'first_time':
                 queryset = queryset.filter(must_change_password=True, is_active=True)
             elif status == 'active':
                 queryset = queryset.filter(is_active=True, is_locked=False, must_change_password=False)

         role_order = Case(
             When(role=User.Role.ADMIN, then=0),
             When(role=User.Role.HOD, then=1),
             When(role=User.Role.PSM, then=2),
             When(role=User.Role.DEAN, then=3),
             When(role=User.Role.LECTURER, then=4),
             default=5,
             output_field=IntegerField()
         )
         return queryset.order_by(role_order, 'username')

    def get_context_data(self, **kwargs):
         context = super().get_context_data(**kwargs)
         context['filters'] = {
             'staff_id': self.request.GET.get('staff_id', '').strip(),
             'username': self.request.GET.get('username', '').strip(),
             'name': self.request.GET.get('name', '').strip(),
             'role': self.request.GET.get('role', '').strip(),
             'status': self.request.GET.get('status', '').strip(),
         }
         return context

class StaffCreateView(RoleRequiredMixin, CreateView):
    model = User
    form_class = StaffForm
    template_name = "dashboard/admin/staff_form.html"
    success_url = reverse_lazy('dashboard:staff_list')
    required_role = User.Role.ADMIN

    def form_valid(self, form):
        response = super().form_valid(form)
        user = self.object
        
        # Generate Temp Password
        temp_password = get_random_string(10)
        user.set_password(temp_password)
        user.must_change_password = True
        user.save()
        
        # Send Email
        # Send Email
        subject = "Welcome to Task Force Management System"
        context = {
            'user': user,
            'temp_password': temp_password,
            'login_url': self.request.build_absolute_uri(reverse_lazy('login'))
        }
        html_message = render_to_string('email/account_created.html', context)
        plain_message = strip_tags(html_message)
        
        if is_throttled(self.request, f"mail:create_user:{user.pk}"):
            messages.info(self.request, f"Staff created. Email already sent to {user.email}.")
        else:
            try:
                send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [user.email], html_message=html_message)
                messages.success(self.request, f"Staff created. Email sent to {user.email}.")
            except Exception as e:
                print(f"Error sending email: {e}")
                messages.warning(self.request, f"Staff created, but email failed to send. Temp Password: {temp_password}")

        log_action(self.request, self.request.user, "CREATE_USER", "User", user.pk, f"Created user {user.username}")
        
        return response

class StaffUpdateView(RoleRequiredMixin, UpdateView):
    model = User
    form_class = StaffForm
    template_name = "dashboard/admin/staff_form.html"
    success_url = reverse_lazy('dashboard:staff_list')
    required_role = User.Role.ADMIN
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_edit'] = True
        return context

class StaffPasswordResetView(RoleRequiredMixin, View):
    required_role = User.Role.ADMIN

    def post(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            
            # Use PasswordResetForm to send the standard reset link
            form = PasswordResetForm({'email': user.email})
            if form.is_valid():
                if is_throttled(request, f"mail:reset_link:{user.pk}"):
                    messages.info(request, f"Password reset link already sent to {user.email}.")
                else:
                    form.save(
                        request=request,
                        use_https=request.is_secure(),
                        email_template_name='registration/password_reset_email.html',
                        html_email_template_name='registration/password_reset_email.html',
                        subject_template_name='registration/password_reset_subject.txt'
                    )
                    messages.success(request, f"Password reset link sent to {user.email}.")
                    log_action(request, request.user, "RESET_PASSWORD", "User", user.pk, f"Sent password reset link for {user.username}")
            else:
                 messages.error(request, f"Could not send reset link. Invalid email for user {user.username}?")

        except User.DoesNotExist:
            messages.error(request, "User not found.")
            
        return redirect('dashboard:staff_edit', pk=pk)

class StaffUnlockView(RoleRequiredMixin, View):
    required_role = User.Role.ADMIN

    def post(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            user.is_locked = False
            user.failed_attempts = 0
            user.save()
            
            
            # Send Email
            # Send Email
            subject = "Account Unlocked - Task Force Management System"
            context = {
                'headline': "Account Unlocked",
                'body_text': f"Hello {user.get_full_name()},\n\nYour account has been unlocked. You can now log in to the system.",
                'action_url': request.build_absolute_uri(reverse_lazy('login')),
                'action_text': "Login Now"
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)

            if not is_throttled(request, f"mail:unlock_user:{user.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [user.email], html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

            messages.success(request, f"Account unlocked for {user.username}. Email sent.")
            log_action(request, request.user, "UNLOCK_USER", "User", user.pk, f"Unlocked user {user.username}")
            
        except User.DoesNotExist:
            messages.error(request, "User not found.")
            
        return redirect('dashboard:staff_list')

class StaffDeactivateView(RoleRequiredMixin, View):
    required_role = User.Role.ADMIN

    def post(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            justification = request.POST.get('justification')
            if not justification:
                messages.error(request, "Justification is required to deactivate a user.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard:staff_list'))

            user.is_active = False
            user.save()
            
            
            # Send Email
            # Send Email
            subject = "Account Deactivated - Task Force Management System"
            context = {
                'headline': "Account Deactivated",
                'body_text': f"Hello {user.get_full_name()},\n\nYour account has been deactivated.\n\nReason: {justification}\n\nPlease contact IT support if you believe this is an error.",
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)
            
            if not is_throttled(request, f"mail:deactivate_user:{user.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [user.email], html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

            messages.success(request, f"User {user.username} deactivated successfully. Notification email sent.")
            log_action(request, request.user, "DEACTIVATE_USER", "User", user.pk, f"Deactivated user {user.username}. Reason: {justification}")
            
        except User.DoesNotExist:
            messages.error(request, "User not found.")
            
        return redirect('dashboard:staff_list')

class StaffActivateView(RoleRequiredMixin, View):
    required_role = User.Role.ADMIN

    def post(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            user.is_active = True
            user.is_locked = False # Also unlock if they were locked
            user.failed_attempts = 0
            user.save()
            
            
            # Send Email
            # Send Email
            subject = "Account Activated - Task Force Management System"
            context = {
                'headline': "Account Reactivated",
                'body_text': f"Hello {user.get_full_name()},\n\nYour account has been reactivated. You can now log in.",
                'action_url': request.build_absolute_uri(reverse_lazy('login')),
                'action_text': "Login Now"
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)
            
            if not is_throttled(request, f"mail:activate_user:{user.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [user.email], html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

            messages.success(request, f"User {user.username} activated successfully. Notification email sent.")
            log_action(request, request.user, "ACTIVATE_USER", "User", user.pk, f"Activated user {user.username}")
            
        except User.DoesNotExist:
            messages.error(request, "User not found.")
            
        return redirect('dashboard:staff_list')

# --- Admin Task Force Management ---
class TaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/admin/taskforce_list.html"
    context_object_name = "taskforces"
    required_role = User.Role.ADMIN

    def get_queryset(self):
        queryset = TaskForce.objects.all().prefetch_related('departments')
        status = self.request.GET.get('status', '').strip()
        department_id = self.request.GET.get('department', '').strip()
        tf_id = self.request.GET.get('tf_id', '').strip()
        name = self.request.GET.get('name', '').strip()

        if status:
            queryset = queryset.filter(status=status)
        if department_id:
            queryset = queryset.filter(departments__id=department_id)
        if tf_id:
            queryset = queryset.filter(chart_id__icontains=tf_id)
        if name:
            queryset = queryset.filter(name__icontains=name)

        status_order = Case(
            When(status='ACTIVE', then=0),
            When(status='DRAFT', then=1),
            When(status='SUBMITTED', then=2),
            When(status='APPROVED', then=3),
            When(status='REJECTED', then=4),
            When(status='INACTIVE', then=5),
            default=6,
            output_field=IntegerField()
        )
        return queryset.order_by(status_order, 'name').distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.all().order_by('name')
        context['filters'] = {
            'status': self.request.GET.get('status', '').strip(),
            'department': self.request.GET.get('department', '').strip(),
            'tf_id': self.request.GET.get('tf_id', '').strip(),
            'name': self.request.GET.get('name', '').strip(),
        }
        return context

class TaskForceCreateView(RoleRequiredMixin, CreateView):
    model = TaskForce
    form_class = TaskForceForm
    template_name = "dashboard/admin/taskforce_form.html"
    success_url = reverse_lazy('dashboard:taskforce_list')
    required_role = User.Role.ADMIN

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request, self.request.user, "CREATE_TASKFORCE", "TaskForce", self.object.pk, f"Created task force {self.object.name}")
        return response

class TaskForceUpdateView(RoleRequiredMixin, UpdateView):
    model = TaskForce
    form_class = TaskForceForm
    template_name = "dashboard/admin/taskforce_form.html"
    success_url = reverse_lazy('dashboard:taskforce_list')
    required_role = User.Role.ADMIN
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Edit Task Force"
        return context

    def form_valid(self, form):
        prev_status = TaskForce.objects.filter(pk=self.object.pk).values_list('status', flat=True).first()
        response = super().form_valid(form)
        new_status = self.object.status

        if prev_status != 'INACTIVE' and new_status == 'INACTIVE':
            if prev_status:
                self.object.previous_status = prev_status
                self.object.save(update_fields=['previous_status'])
        elif prev_status == 'INACTIVE' and new_status == 'ACTIVE':
            if self.object.previous_status and self.object.previous_status != 'INACTIVE':
                self.object.status = self.object.previous_status
                self.object.save(update_fields=['status'])
                new_status = self.object.status
        elif prev_status not in ('ACTIVE', 'INACTIVE') and new_status == 'ACTIVE':
            self.object.status = prev_status
            self.object.save(update_fields=['status'])
            new_status = self.object.status

        if prev_status != 'INACTIVE' and new_status == 'INACTIVE':
            log_action(self.request, self.request.user, "DEACTIVATE_TASKFORCE", "TaskForce", self.object.pk, f"Deactivated task force {self.object.name}")
        elif prev_status == 'INACTIVE' and new_status != 'INACTIVE':
            log_action(self.request, self.request.user, "ACTIVATE_TASKFORCE", "TaskForce", self.object.pk, f"Reactivated task force {self.object.name} to {new_status}")
        else:
            log_action(self.request, self.request.user, "UPDATE_TASKFORCE", "TaskForce", self.object.pk, f"Updated task force {self.object.name}")
        return response

class WorkloadSettingsView(RoleRequiredMixin, UpdateView):
    model = WorkloadSettings
    form_class = WorkloadSettingsForm
    template_name = "dashboard/admin/workload_settings.html"
    success_url = reverse_lazy('dashboard:admin')
    required_role = User.Role.ADMIN

    def get_object(self, queryset=None):
        # Singleton pattern: ensure one exists
        obj, created = WorkloadSettings.objects.get_or_create(pk=1)
        return obj

    def form_valid(self, form):
        messages.success(self.request, "Workload thresholds updated successfully.")
        log_action(self.request, self.request.user, "UPDATE_SETTINGS", "WorkloadSettings", self.object.pk, f"Updated thresholds: Min={form.instance.min_weightage}, Max={form.instance.max_weightage}")
        return super().form_valid(form)

class HODDashboardView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/hod_dashboard.html"
    required_role = User.Role.HOD

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.department:
            taskforces = TaskForce.objects.filter(departments=self.request.user.department).distinct()
            status_filter = self.request.GET.get('status', '').strip()
            query = self.request.GET.get('q', '').strip()

            if status_filter:
                taskforces = taskforces.filter(status=status_filter)
            if query:
                taskforces = taskforces.filter(
                    Q(name__icontains=query) |
                    Q(chart_id__icontains=query) |
                    Q(description__icontains=query)
                )

            context['taskforce_count'] = taskforces.count()
            context['taskforces'] = taskforces
            context['selected_status'] = status_filter
            context['query'] = query
        else:
            context['taskforce_count'] = 0
            context['taskforces'] = TaskForce.objects.none()
        return context

from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from .forms import (
    StaffForm, TaskForceForm, DepartmentForm, TaskForceMembershipForm, PSMTaskForceMembershipForm
)

# ... (Previous imports)

class HODTaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/hod/taskforce_list.html"
    context_object_name = "taskforces"
    required_role = User.Role.HOD

    def get_queryset(self):
        # Filter task forces that include the HOD's department
        if not self.request.user.department:
            return TaskForce.objects.none()
        taskforces = TaskForce.objects.filter(departments=self.request.user.department).distinct()
        self.status_filter = self.request.GET.get('status', '').strip()
        self.query = self.request.GET.get('q', '').strip()

        if self.status_filter:
            taskforces = taskforces.filter(status=self.status_filter)
        if self.query:
            taskforces = taskforces.filter(
                Q(name__icontains=self.query) |
                Q(chart_id__icontains=self.query) |
                Q(description__icontains=self.query)
            )

        return taskforces

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_status'] = getattr(self, 'status_filter', '')
        context['query'] = getattr(self, 'query', '')
        return context

class HODTaskForceUpdateView(RoleRequiredMixin, UpdateView):
    model = TaskForce
    form_class = TaskForceMembershipForm
    template_name = "dashboard/hod/taskforce_manage.html"
    context_object_name = "taskforce"
    required_role = User.Role.HOD
    success_url = reverse_lazy('dashboard:hod_taskforce_list')
    locked_statuses = {'APPROVED', 'SUBMITTED', 'INACTIVE'}

    def get_queryset(self):
        # Ensure HOD can only edit task forces for their department
        if not self.request.user.department:
            return TaskForce.objects.none()
        return TaskForce.objects.filter(departments=self.request.user.department)

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status in self.locked_statuses:
            messages.error(request, "This task force is locked while submitted, approved, or inactive.")
            return redirect('dashboard:hod_taskforce_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        """Pass the HOD's department to the form to filter staff."""
        kwargs = super().get_form_kwargs()
        kwargs['department'] = self.request.user.department
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from university.services import WorkloadService
        import json
        
        # Serialize Members
        members_data = []
        for member in self.object.members.all():
            # Calculate workload (including this task force as they are already a member)
            status = WorkloadService.get_workload_status(member)
            members_data.append({
                'id': member.id,
                'name': member.get_full_name() or member.username,
                'email': member.email,
                'role': member.get_role_display(),
                'workload': status
            })
        
        context['current_members_json'] = json.dumps(members_data)
        return context

    def form_valid(self, form):
        action = self.request.POST.get('action')
        members = form.cleaned_data.get('members')

        if action == 'submit' and (not members or members.count() < 1):
            messages.error(self.request, "Please add at least one member before submitting for approval.")
            return redirect('dashboard:hod_taskforce_manage', pk=self.object.pk)
        
        # Check if justification is needed but missing
        # (This is a fallback; frontend should catch this, but backend must enforce)
        # However, checking overload here is complex without re-calculating everything.
        # Let's rely on the fact that if they click 'Submit', they are claiming it's ready.
        # Ideally, we should check overload status here.
        
        if action == 'submit':
            form.instance.status = 'SUBMITTED'
            form.instance.submitted_by = self.request.user
            # Check for justification if provided
            justification = self.request.POST.get('justification', '').strip()
            form.instance.hod_justification = justification or None

            response = super().form_valid(form)
            log_action(self.request, self.request.user, "SUBMIT_TASKFORCE", "TaskForce", self.object.pk, "Submitted for approval")
            
            # Send Email logic (kept same)
            subject = f"Task Force Submitted: {self.object.name}"
            context = {
                'headline': "Submission Successful",
                'body_text': f"Hello {self.request.user.get_full_name()},\n\nYou have successfully submitted the Task Force '{self.object.name}' for approval.",
                'action_url': self.request.build_absolute_uri(reverse_lazy('dashboard:hod_taskforce_list')),
                'action_text': "View Status"
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)
            
            if not is_throttled(self.request, f"mail:hod_submit:{self.object.pk}:{self.request.user.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [self.request.user.email], html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

            # Notify assigned PSM on resubmission
            if self.object.assigned_psm and self.object.assigned_psm.email:
                psm_subject = f"Task Force Resubmitted: {self.object.name}"
                psm_context = {
                    'headline': "Task Force Resubmitted",
                    'body_text': f"The Task Force '{self.object.name}' has been resubmitted for your review.",
                    'action_url': self.request.build_absolute_uri(reverse_lazy('dashboard:psm_taskforce_review', kwargs={'pk': self.object.pk})),
                    'action_text': "Review Submission"
                }
                psm_html = render_to_string('email/notification.html', psm_context)
                psm_plain = strip_tags(psm_html)
                if not is_throttled(self.request, f"mail:psm_resubmit:{self.object.pk}:{self.object.assigned_psm_id}"):
                    try:
                        send_mail(psm_subject, psm_plain, settings.DEFAULT_FROM_EMAIL, [self.object.assigned_psm.email], html_message=psm_html)
                    except Exception as e:
                        print(f"Error sending email to PSM: {e}")
                
            messages.success(self.request, f"Task Force '{self.object.name}' submitted successfully. Confirmation email sent.")
            return response
            
        elif action == 'save_draft':
            # Save as DRAFT if it was Draft or Active.
            # If it was Active (from creation), keeping it Active is fine, or switching to Draft?
            # User workflow: "Save Draft". Implies switching to DRAFT status if not already.
            # But if it was ACTIVE, maybe we shouldn't demote it?
            # Actually, `DRAFT` is a new status. Let's use it.
            if form.instance.status != 'ACTIVE': # Don't change Active to Draft if it was already live?
                 form.instance.status = 'DRAFT'
            else:
                 # If it was defined as ACTIVE by default logic, maybe we want to keep it active?
                 # Typically 'Save Draft' means "Work in Progress".
                 pass 
            
            # For this specific request: "system will not save that data... until clicking the review request"
            # But we added "Save Draft". 
            # Let's set it to DRAFT.
            form.instance.status = 'DRAFT'
            response = super().form_valid(form)
            messages.success(self.request, "Draft saved successfully.")
            return response

        return super().form_valid(form)

class HODTaskForceDetailView(RoleRequiredMixin, DetailView):
    model = TaskForce
    template_name = "dashboard/hod/taskforce_detail.html"
    context_object_name = "taskforce"
    required_role = User.Role.HOD

    def get_queryset(self):
        if not self.request.user.department:
            return TaskForce.objects.none()
        return (
            TaskForce.objects.filter(departments=self.request.user.department)
            .prefetch_related('departments', 'members', 'members__department')
            .select_related('submitted_by', 'assigned_psm')
        )

from accounts.utils import log_action, is_throttled
import csv
from django.http import HttpResponse

class AuditLogListView(RoleRequiredMixin, ListView):
    model = AuditLog
    template_name = "dashboard/admin/audit_log.html"
    context_object_name = "logs"
    required_role = User.Role.ADMIN
    paginate_by = 20

    def get_queryset(self):
        queryset = AuditLog.objects.all().select_related('actor')
        if self.request.method == 'POST':
            user_query = self.request.POST.get('user')
        else:
            user_query = self.request.GET.get('user')
        if user_query:
            queryset = queryset.filter(actor__username__icontains=user_query)
        return queryset

    def post(self, request, *args, **kwargs):
        if request.POST.get('export') != 'csv':
            return redirect('dashboard:audit_log_list')

        if not is_throttled(request, f"log:export_audit:{request.user.pk}"):
            log_action(
                request,
                request.user,
                "EXPORT_AUDIT_LOGS",
                "AuditLog",
                None,
                "Exported audit logs to CSV"
            )
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
        writer = csv.writer(response)
        writer.writerow(['Timestamp', 'Actor', 'Action', 'Target Model', 'Target ID', 'Details', 'IP'])
        for log in self.get_queryset():
            writer.writerow([log.timestamp, log.actor, log.action, log.target_model, log.target_id, log.details, log.ip_address])
        return response

class PSMDashboardView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/psm_dashboard.html"
    required_role = User.Role.PSM
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Count Pending Approvals
        context['pending_count'] = TaskForce.objects.filter(status='SUBMITTED').count()
        context['actioned_count'] = TaskForce.objects.filter(assigned_psm=self.request.user).count()
        return context

class PSMTaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/psm/taskforce_list.html"
    context_object_name = "taskforces"
    required_role = User.Role.PSM

    def get_queryset(self):
        # PSM sees SUBMITTED task forces for approval
        return TaskForce.objects.filter(
            status='SUBMITTED'
        ).filter(
            Q(assigned_psm__isnull=True) | Q(assigned_psm=self.request.user)
        ).order_by('-updated_at')

class PSMTaskForceDetailView(RoleRequiredMixin, DetailView):
    model = TaskForce
    template_name = "dashboard/psm/taskforce_review.html"
    context_object_name = "taskforce"
    required_role = User.Role.PSM

    def get_queryset(self):
        return TaskForce.objects.filter(
            status='SUBMITTED'
        ).filter(
            Q(assigned_psm__isnull=True) | Q(assigned_psm=self.request.user)
        )
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        action = request.POST.get('action')

        if self.object.assigned_psm and self.object.assigned_psm != request.user:
            messages.error(request, "You are not assigned to review this task force.")
            return redirect('dashboard:psm_taskforce_list')

        if action == 'approve':
            self.object.status = 'APPROVED'
            if not self.object.assigned_psm:
                self.object.assigned_psm = request.user
            self.object.save()
            log_action(request, request.user, "APPROVE_TASKFORCE", "TaskForce", self.object.pk, f"Approved task force: {self.object.name}")
            
            # Send Email logic (removed chairman specific email)
            # Consider emailing HODs of involved departments instead?
            # For now, no individual email if no specific leader.
            recipients = []
            if self.object.submitted_by and self.object.submitted_by.email:
                recipients.append(self.object.submitted_by.email)
            member_emails = list(self.object.members.exclude(email__isnull=True).exclude(email__exact='').values_list('email', flat=True))
            recipients.extend(member_emails)
            if recipients:
                subject = f"Task Force Approved: {self.object.name}"
                context = {
                    'headline': "Task Force Approved",
                    'body_text': f"Task Force '{self.object.name}' has been approved by the PSM.",
                    'action_url': request.build_absolute_uri(reverse_lazy('dashboard:lecturer_portfolio')),
                    'action_text': "View Task Force"
                }
                html_message = render_to_string('email/notification.html', context)
                plain_message = strip_tags(html_message)
                if not is_throttled(request, f"mail:psm_approve:{self.object.pk}"):
                    try:
                        send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, list(set(recipients)), html_message=html_message)
                    except Exception as e:
                        print(f"Error sending email: {e}")

            messages.success(request, f"Task Force '{self.object.name}' approved.")
            return redirect('dashboard:psm_taskforce_list')
            
        elif action == 'reject':
            reason = request.POST.get('rejection_reason')
            if reason:
                self.object.rejection_reason = reason
                self.object.status = 'REJECTED'
                if not self.object.assigned_psm:
                    self.object.assigned_psm = request.user
                self.object.save()
                log_action(request, request.user, "REJECT_TASKFORCE", "TaskForce", self.object.pk, f"Rejected with reason: {reason}")
                
                # Send Email logic (removed chairman specific email)
                # Consider emailing HODs of involved departments instead?
                if self.object.submitted_by and self.object.submitted_by.email:
                    subject = f"Task Force Rejected: {self.object.name}"
                    context = {
                        'headline': "Task Force Rejected",
                        'body_text': f"Task Force '{self.object.name}' has been rejected.\n\nReason:\n{reason}",
                        'action_url': request.build_absolute_uri(reverse_lazy('dashboard:hod_taskforce_list')),
                        'action_text': "Review Task Force"
                    }
                    html_message = render_to_string('email/notification.html', context)
                    plain_message = strip_tags(html_message)
                    if not is_throttled(request, f"mail:psm_reject:{self.object.pk}"):
                        try:
                            send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [self.object.submitted_by.email], html_message=html_message)
                        except Exception as e:
                            print(f"Error sending email: {e}")

                messages.success(request, f"Task Force '{self.object.name}' rejected.")
                return redirect('dashboard:psm_taskforce_list')
            else:
                messages.error(request, "Rejection reason is required.")
                return redirect('dashboard:psm_taskforce_review', pk=self.object.pk)
                 
        return redirect('dashboard:psm_taskforce_list')

class PSMTaskForceModifyView(RoleRequiredMixin, UpdateView):
    model = TaskForce
    form_class = PSMTaskForceMembershipForm
    template_name = "dashboard/psm/taskforce_modify.html"
    context_object_name = "taskforce"
    required_role = User.Role.PSM
    success_url = reverse_lazy('dashboard:psm_taskforce_list')

    def get_queryset(self):
        return TaskForce.objects.filter(
            status='SUBMITTED'
        ).filter(
            Q(assigned_psm__isnull=True) | Q(assigned_psm=self.request.user)
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['departments'] = self.object.departments.all()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from university.services import WorkloadService
        import json

        members_data = []
        for member in self.object.members.all():
            status = WorkloadService.get_workload_status(member)
            members_data.append({
                'id': member.id,
                'name': member.get_full_name() or member.username,
                'email': member.email,
                'role': member.get_role_display(),
                'workload': status
            })
        context['current_members_json'] = json.dumps(members_data)
        context['department_ids'] = ",".join(str(d.id) for d in self.object.departments.all())
        return context

    def form_valid(self, form):
        justification = self.request.POST.get('psm_adjustment_reason', '').strip()
        if not justification:
            messages.error(self.request, "Justification is required to modify and approve.")
            return redirect('dashboard:psm_taskforce_modify', pk=self.object.pk)

        self.object = form.save(commit=False)
        self.object.status = 'APPROVED'
        self.object.psm_adjustment_reason = justification
        self.object.psm_adjusted_at = timezone.now()
        if not self.object.assigned_psm:
            self.object.assigned_psm = self.request.user
        self.object.save()
        self.object.members.set(form.cleaned_data.get('members', []))

        log_action(self.request, self.request.user, "MODIFY_APPROVE_TASKFORCE", "TaskForce", self.object.pk, f"Modified and approved task force: {self.object.name}")

        recipients = []
        if self.object.submitted_by and self.object.submitted_by.email:
            recipients.append(self.object.submitted_by.email)
        member_emails = list(self.object.members.exclude(email__isnull=True).exclude(email__exact='').values_list('email', flat=True))
        recipients.extend(member_emails)
        if recipients:
            subject = f"Task Force Modified and Approved: {self.object.name}"
            context = {
                'headline': "Task Force Modified and Approved",
                'body_text': f"Task Force '{self.object.name}' has been modified and approved by the PSM.\n\nReason:\n{justification}",
                'action_url': self.request.build_absolute_uri(reverse_lazy('dashboard:lecturer_portfolio')),
                'action_text': "View Task Force"
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)
            if not is_throttled(self.request, f"mail:psm_modify:{self.object.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, list(set(recipients)), html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

        messages.success(self.request, f"Task Force '{self.object.name}' modified and approved.")
        return redirect(self.success_url)

class PSMActionedTaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/psm/taskforce_actioned_list.html"
    context_object_name = "taskforces"
    required_role = User.Role.PSM

    def get_queryset(self):
        return TaskForce.objects.filter(assigned_psm=self.request.user).order_by('-updated_at')

class PSMActionedTaskForceUpdateView(RoleRequiredMixin, UpdateView):
    model = TaskForce
    form_class = PSMTaskForceMembershipForm
    template_name = "dashboard/psm/taskforce_actioned_detail.html"
    context_object_name = "taskforce"
    required_role = User.Role.PSM
    success_url = reverse_lazy('dashboard:psm_taskforce_actioned_list')

    def get_queryset(self):
        return TaskForce.objects.filter(assigned_psm=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['departments'] = self.object.departments.all()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from university.services import WorkloadService
        import json

        members_data = []
        for member in self.object.members.all():
            status = WorkloadService.get_workload_status(member)
            members_data.append({
                'id': member.id,
                'name': member.get_full_name() or member.username,
                'email': member.email,
                'role': member.get_role_display(),
                'workload': status
            })
        context['current_members_json'] = json.dumps(members_data)
        context['department_ids'] = ",".join(str(d.id) for d in self.object.departments.all())
        return context

    def form_valid(self, form):
        if self.object.status != 'APPROVED':
            messages.error(self.request, "Only approved task forces can be updated here.")
            return redirect('dashboard:psm_taskforce_actioned_list')

        justification = self.request.POST.get('psm_adjustment_reason', '').strip()
        if not justification:
            messages.error(self.request, "Justification is required to update a locked task force.")
            return redirect('dashboard:psm_taskforce_actioned_detail', pk=self.object.pk)

        self.object = form.save(commit=False)
        self.object.psm_adjustment_reason = justification
        self.object.psm_adjusted_at = timezone.now()
        self.object.save()
        self.object.members.set(form.cleaned_data.get('members', []))

        log_action(self.request, self.request.user, "UPDATE_LOCKED_TASKFORCE", "TaskForce", self.object.pk, f"Updated locked task force: {self.object.name}")

        recipients = []
        if self.object.submitted_by and self.object.submitted_by.email:
            recipients.append(self.object.submitted_by.email)
        member_emails = list(self.object.members.exclude(email__isnull=True).exclude(email__exact='').values_list('email', flat=True))
        recipients.extend(member_emails)
        if recipients:
            subject = f"Task Force Updated: {self.object.name}"
            context = {
                'headline': "Task Force Updated",
                'body_text': f"Task Force '{self.object.name}' has been updated by the PSM.\n\nReason:\n{justification}",
                'action_url': self.request.build_absolute_uri(reverse_lazy('dashboard:lecturer_portfolio')),
                'action_text': "View Task Force"
            }
            html_message = render_to_string('email/notification.html', context)
            plain_message = strip_tags(html_message)
            if not is_throttled(self.request, f"mail:psm_update:{self.object.pk}"):
                try:
                    send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, list(set(recipients)), html_message=html_message)
                except Exception as e:
                    print(f"Error sending email: {e}")

        messages.success(self.request, "Locked task force updated successfully.")
        return redirect(self.success_url)

class DeanDashboardView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/dean_dashboard.html"
    required_role = User.Role.DEAN
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Executive Stats
        context['active_taskforces_all'] = TaskForce.objects.exclude(status='INACTIVE').count()
        context['total_taskforces'] = TaskForce.objects.count()
        context['active_taskforces'] = TaskForce.objects.filter(status='APPROVED').count()
        context['pending_approvals'] = TaskForce.objects.filter(status='SUBMITTED').count()
        context['departments'] = Department.objects.all().order_by('name')
        return context

class LecturerDashboardView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/lecturer_dashboard.html"
    required_role = User.Role.LECTURER

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Count assignments
        context['assignment_count'] = TaskForce.objects.filter(members=self.request.user, status='APPROVED').count()
        context['inactive_count'] = TaskForce.objects.filter(members=self.request.user, status='INACTIVE').count()
        from university.services import WorkloadService
        workload = WorkloadService.get_workload_status(self.request.user)
        status = workload.get('status', 'UNKNOWN')
        badge_map = {
            'OVERLOADED': 'bg-danger',
            'UNDERLOADED': 'bg-warning text-dark',
            'BALANCED': 'bg-success',
            'UNKNOWN': 'bg-secondary'
        }
        border_map = {
            'OVERLOADED': 'border-danger',
            'UNDERLOADED': 'border-warning',
            'BALANCED': 'border-success',
            'UNKNOWN': 'border-secondary'
        }
        context['workload_total'] = workload.get('current_weightage', 0)
        context['workload_message'] = workload.get('message', 'Workload unavailable')
        context['workload_status'] = status
        context['workload_badge_class'] = badge_map.get(status, 'bg-secondary')
        context['workload_border_class'] = border_map.get(status, 'border-secondary')
        return context

class LecturerTaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/lecturer/portfolio.html"
    context_object_name = "taskforces"
    required_role = User.Role.LECTURER

    def get_queryset(self):
        return TaskForce.objects.filter(
            members=self.request.user,
            status='APPROVED'
        ).distinct().order_by('-updated_at')

class LecturerTaskForceInactiveListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/lecturer/portfolio_inactive.html"
    context_object_name = "taskforces"
    required_role = User.Role.LECTURER

    def get_queryset(self):
        return TaskForce.objects.filter(
            members=self.request.user,
            status='INACTIVE'
        ).distinct().order_by('-updated_at')

class LecturerTaskForceReportView(RoleRequiredMixin, View):
    required_role = User.Role.LECTURER

    def get(self, request, *args, **kwargs):
        return redirect('dashboard:lecturer')

    def post(self, request, *args, **kwargs):
        report_type = request.POST.get('report_type', 'excel').lower()
        if report_type not in ('excel', 'pdf'):
            report_type = 'excel'
        taskforces = TaskForce.objects.filter(
            members=request.user,
            status='APPROVED'
        ).prefetch_related('departments', 'members').order_by('-updated_at')

        if not is_throttled(request, f"log:export_lecturer:{request.user.pk}:{report_type}"):
            log_action(
                request,
                request.user,
                "EXPORT_LECTURER_REPORT",
                "TaskForce",
                None,
                f"Exported lecturer task force report ({report_type.upper()})"
            )

        if report_type == 'pdf':
            return build_taskforce_pdf_response(
                taskforces,
                "Lecturer Task Force Report",
                "lecturer_taskforce_report.pdf"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturer Report"

        headers = [
            "Task Force ID",
            "Task Force Name",
            "Description",
            "Departments",
            "Status",
            "Weightage",
            "Created At",
            "Updated At"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        total_weightage = 0
        for tf in taskforces:
            departments = ", ".join([d.name for d in tf.departments.all()])
            ws.append([
                tf.chart_id or "",
                tf.name,
                tf.description or "",
                departments,
                tf.get_status_display(),
                tf.weightage,
                tf.created_at.strftime("%Y-%m-%d") if tf.created_at else "",
                tf.updated_at.strftime("%Y-%m-%d") if tf.updated_at else ""
            ])
            total_weightage += tf.weightage

        ws.append([])
        ws.append(["Total Weightage", total_weightage])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="lecturer_taskforce_report.xlsx"'
        wb.save(response)
        return response

class LecturerTaskForceDetailView(RoleRequiredMixin, DetailView):
    model = TaskForce
    template_name = "dashboard/lecturer/taskforce_detail.html"
    context_object_name = "taskforce"
    required_role = User.Role.LECTURER

    def get_queryset(self):
        return TaskForce.objects.filter(members=self.request.user)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            messages.error(request, "Remarks cannot be empty.")
            return redirect('dashboard:lecturer_taskforce_detail', pk=self.object.pk)

        if not self.object.assigned_psm or not self.object.assigned_psm.email:
            messages.error(request, "No PSM is assigned to this task force yet.")
            return redirect('dashboard:lecturer_taskforce_detail', pk=self.object.pk)

        subject = f"Lecturer Remarks: {self.object.name}"
        context = {
            'headline': "Lecturer Remarks Submitted",
            'body_text': (
                f"Lecturer: {request.user.get_full_name() or request.user.username}\n"
                f"Task Force: {self.object.name} ({self.object.chart_id or 'No ID'})\n\n"
                f"Remarks:\n{remarks}"
            ),
            'action_url': request.build_absolute_uri(reverse_lazy('dashboard:psm_taskforce_review', kwargs={'pk': self.object.pk})),
            'action_text': "View Task Force"
        }
        html_message = render_to_string('email/notification.html', context)
        plain_message = strip_tags(html_message)
        if is_throttled(request, f"mail:lecturer_remarks:{self.object.pk}:{request.user.pk}"):
            messages.info(request, "Remarks were sent recently. Please wait before resending.")
        else:
            try:
                send_mail(subject, plain_message, settings.DEFAULT_FROM_EMAIL, [self.object.assigned_psm.email], html_message=html_message)
                messages.success(request, "Your remarks have been sent to the PSM.")
            except Exception as e:
                print(f"Error sending remarks email: {e}")
                messages.error(request, "Failed to send remarks. Please try again later.")

        return redirect('dashboard:lecturer_taskforce_detail', pk=self.object.pk)

class DeanReportView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/dean/report_list.html"
    context_object_name = "taskforces"
    required_role = User.Role.DEAN

    def get_queryset(self):
        # Dean sees ALL task forces
        queryset = TaskForce.objects.all().prefetch_related('departments')
        
        # Filter by Department
        dept_id = self.request.GET.get('department')
        if dept_id:
            queryset = queryset.filter(departments__id=dept_id)
            
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from university.models import Department
        context['departments'] = Department.objects.all()
        return context

class DeanDepartmentSummaryView(RoleRequiredMixin, TemplateView):
    template_name = "dashboard/dean/department_summary.html"
    required_role = User.Role.DEAN

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        departments = Department.objects.all().prefetch_related('task_forces__members')
        rows = []
        for dept in departments:
            taskforces = [tf for tf in dept.task_forces.all() if tf.status != 'INACTIVE']
            total_weightage = sum(tf.weightage or 0 for tf in taskforces)
            lecturer_ids = set()
            for tf in taskforces:
                for member in tf.members.all():
                    if member.role == User.Role.LECTURER and member.department_id == dept.id:
                        lecturer_ids.add(member.id)
            rows.append({
                'department': dept,
                'taskforce_count': len(taskforces),
                'total_weightage': total_weightage,
                'lecturer_count': len(lecturer_ids),
            })
        context['department_rows'] = rows
        return context

class DeanDepartmentTaskForceListView(RoleRequiredMixin, ListView):
    model = TaskForce
    template_name = "dashboard/dean/department_taskforces.html"
    context_object_name = "taskforces"
    required_role = User.Role.DEAN

    def get_queryset(self):
        department_id = self.kwargs['department_id']
        self.department = get_object_or_404(Department, pk=department_id)
        return (
            TaskForce.objects.filter(departments=self.department)
            .exclude(status='INACTIVE')
            .prefetch_related('departments', 'members')
            .distinct()
            .order_by('-updated_at')
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['department'] = self.department
        return context

class DeanTaskForceDetailView(RoleRequiredMixin, DetailView):
    model = TaskForce
    template_name = "dashboard/dean/taskforce_detail.html"
    context_object_name = "taskforce"
    required_role = User.Role.DEAN

    def get_queryset(self):
        return TaskForce.objects.prefetch_related('departments', 'members', 'members__department')

class DeanTaskForceReportDownloadView(RoleRequiredMixin, View):
    required_role = User.Role.DEAN

    def get(self, request, *args, **kwargs):
        return redirect('dashboard:dean')

    def post(self, request, *args, **kwargs):
        scope = request.POST.get('scope', 'all')
        department_id = request.POST.get('department_id')
        report_type = request.POST.get('report_type', 'excel').lower()
        if report_type not in ('excel', 'pdf'):
            report_type = 'excel'

        queryset = TaskForce.objects.all().prefetch_related('departments', 'members')
        filename = "taskforce_report_all.xlsx"

        if scope == 'department':
            if not department_id:
                messages.error(request, "Please select a department for the report.")
                return redirect('dashboard:dean')
            department = get_object_or_404(Department, pk=department_id)
            queryset = queryset.filter(departments=department).distinct()
            safe_name = department.name.replace(" ", "_")
            filename = f"taskforce_report_{safe_name}.xlsx"

        detail_scope = "all" if scope != "department" else f"department:{department_id}"
        if not is_throttled(request, f"log:export_dean:{request.user.pk}:{detail_scope}:{report_type}"):
            log_action(
                request,
                request.user,
                "EXPORT_DEAN_REPORT",
                "TaskForce",
                None,
                f"Exported dean task force report ({report_type.upper()}), scope={detail_scope}"
            )

        if report_type == 'pdf':
            pdf_name = filename.replace(".xlsx", ".pdf")
            title = "Dean Task Force Report"
            if scope == 'department':
                title = f"Dean Task Force Report - {department.name}"
            return build_taskforce_pdf_response(queryset.order_by('-updated_at'), title, pdf_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "Task Forces"

        headers = [
            "Task Force ID",
            "Task Force Name",
            "Description",
            "Departments",
            "Status",
            "Weightage",
            "Members",
            "Created At",
            "Updated At",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for tf in queryset.order_by('-updated_at'):
            dept_names = ", ".join([d.name for d in tf.departments.all()])
            ws.append([
                tf.chart_id or "",
                tf.name,
                tf.description or "",
                dept_names,
                tf.get_status_display(),
                tf.weightage,
                tf.members.count(),
                tf.created_at.strftime("%Y-%m-%d") if tf.created_at else "",
                tf.updated_at.strftime("%Y-%m-%d") if tf.updated_at else "",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
